import tkinter as tk
from tkinter import messagebox
import sys
import os
import subprocess
from tkinter import ttk
from tkinter import filedialog
from pathlib import Path
import json
import shutil
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from tkinter import colorchooser
import win32print
import win32ui
from PIL import Image, ImageWin

#Core Assignments
SETTINGS = Path(__file__).with_name("settings.json")
with open(SETTINGS, 'r') as f:
    SETTINGS_DATA = json.load(f)
icon_dir = Path(__file__).parent / "icons"
ICON_PATH = icon_dir / "icon.png"

load = tk.Tk()
load.title("ODIN - Loading...")
load.geometry("600x600")
icon = tk.PhotoImage(file=str(ICON_PATH))
load.iconphoto(False, icon)
logo = tk.PhotoImage(file=str(icon_dir / "logo.png"))
logo_label = ttk.Label(load, image=logo)
logo_label.pack(pady=10)
ttk.Label(load, text="ODIN is loading, please wait...", font=('TkDefaultFont', 12)).pack(expand=True)
load_time = 3000  # milliseconds
load.update()
load.after(load_time, load.destroy)
load.mainloop()

#Main Window

root = tk.Tk()
root.title(f"{SETTINGS_DATA.get('app_name', ' ')} - Version {SETTINGS_DATA.get('version', 'Version Unknown')}")
icon = tk.PhotoImage(file=str(ICON_PATH))
root.iconphoto(False, icon)
root.geometry("1200x800")
root.update_idletasks()

#Core Functions

def convert_report(xls_path):
    """Convert .xls file to .xlsx format for openpyxl compatibility."""
    try:
        # Read the .xls file with pandas
        df = pd.read_excel(xls_path, engine='xlrd')
        
        # Create .xlsx path
        xlsx_path = Path(xls_path).with_suffix('.xlsx')
        
        # Write to .xlsx format
        df.to_excel(xlsx_path, index=False, engine='openpyxl')
        
        return xlsx_path
    except Exception as e:
        messagebox.showerror("Conversion Error", f"Failed to convert {xls_path}: {e}")
        return None

def process_report(xlsx_path, status):
    """Process .xlsx report into JSON, deleting column K and renaming column L to 'Customer'."""
    try:
        original_xls_path = None
        
        # Convert .xls to .xlsx if necessary
        if str(xlsx_path).endswith('.xls'):
            original_xls_path = xlsx_path
            xlsx_path = convert_report(xlsx_path)
            if xlsx_path is None:
                return
        
        # Load the workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Delete column K (11th column)
        ws.delete_cols(11)
        
        # Rename column L (which is now column K after deletion) to 'Customer'
        ws.cell(row=1, column=11).value = 'Customer'
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Get current timestamp for Last Update
        current_time = datetime.now().isoformat()
        today = datetime.now().date()
        
        # Process data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip row if column A (index 0) has less than 3 characters
            if row[0] is None or len(str(row[0])) < 3:
                continue
            
            row_data = {}
            for idx, (header, value) in enumerate(zip(headers, row)):
                # Convert datetime objects to ISO format string (column J is index 9)
                if idx == 9 and isinstance(value, datetime):
                    row_data[header] = value.isoformat()
                else:
                    row_data[header] = value
            
            # Add status, priority, and last update to each entry
            row_data['Status'] = status
            row_data['Priority'] = False
            row_data['Last Update'] = current_time
            data.append(row_data)
        
        # Load existing database.json
        db_path = Path(__file__).parent / "database.json"
        existing_data = []
        if db_path.exists():
            try:
                with open(db_path, 'r') as f:
                    existing_data = json.load(f)
            except Exception:
                existing_data = []
        
        # Save new data to temporary database
        temp_db_path = Path(__file__).parent / "database_temp.json"
        with open(temp_db_path, 'w') as f:
            json.dump(data, f, indent=4)
        
        # Compare old and new data for telemetrics
        old_job_nos = {entry.get('Job No') for entry in existing_data if entry.get('Job No')}
        new_job_nos = {entry.get('Job No') for entry in data if entry.get('Job No')}
        
        jobs_removed = old_job_nos - new_job_nos
        jobs_added = new_job_nos - old_job_nos
        jobs_retained = old_job_nos & new_job_nos
        
        # Calculate turnaround time for removed jobs
        removed_jobs_details = []
        for job_no in jobs_removed:
            # Find the job in existing_data
            for entry in existing_data:
                if entry.get('Job No') == job_no:
                    rec_date = entry.get('Rec_Date', '')
                    turnaround_days = None
                    
                    if rec_date:
                        try:
                            if isinstance(rec_date, str):
                                rec_date_obj = datetime.fromisoformat(rec_date).date()
                            else:
                                rec_date_obj = rec_date.date() if hasattr(rec_date, 'date') else rec_date
                            
                            turnaround_days = (today - rec_date_obj).days
                        except Exception:
                            pass
                    
                    removed_jobs_details.append({
                        'job_no': job_no,
                        'customer': entry.get('Customer', ''),
                        'turnaround_days': turnaround_days
                    })
                    break
        
        # Get customer names for added jobs
        added_jobs_details = []
        for job_no in jobs_added:
            for entry in data:
                if entry.get('Job No') == job_no:
                    added_jobs_details.append({
                        'job_no': job_no,
                        'customer': entry.get('Customer', '')
                    })
                    break
        
        # Create log entry
        log_entry = {
            'timestamp': current_time,
            'date': datetime.now().strftime('%Y-%m-%d'),
            'report_type': status,
            'total_jobs_before': len(existing_data),
            'total_jobs_after': len(data),
            'jobs_added': len(jobs_added),
            'jobs_removed': len(jobs_removed),
            'jobs_retained': len(jobs_retained),
            'removed_jobs': removed_jobs_details,
            'added_jobs': added_jobs_details
        }
        
        # Load or create log.json
        log_path = Path(__file__).parent / "log.json"
        log_data = []
        if log_path.exists():
            try:
                with open(log_path, 'r') as f:
                    log_data = json.load(f)
            except Exception:
                log_data = []
        
        # Append new log entry
        log_data.append(log_entry)
        
        # Save log
        with open(log_path, 'w') as f:
            json.dump(log_data, f, indent=4)
        
        # Merge new data with existing data, preserving metadata from existing entries
        merged_data = []
        for new_entry in data:
            job_no = new_entry.get('Job No')
            
            # Find matching entry in existing data
            existing_entry = None
            for old_entry in existing_data:
                if old_entry.get('Job No') == job_no:
                    existing_entry = old_entry
                    break
            
            # Preserve Priority, Assigned, and original Last Update if job existed
            if existing_entry:
                new_entry['Priority'] = existing_entry.get('Priority', False)
                new_entry['Assigned'] = existing_entry.get('Assigned', '')
                # Keep existing Last Update if no changes to core fields
                if existing_entry.get('Status') == new_entry.get('Status'):
                    new_entry['Last Update'] = existing_entry.get('Last Update', current_time)
            
            merged_data.append(new_entry)
        
        # Save merged data to database.json
        with open(db_path, 'w') as f:
            json.dump(merged_data, f, indent=4)
        
        # Delete temporary files
        try:
            if original_xls_path and Path(original_xls_path).exists():
                Path(original_xls_path).unlink()
            if Path(xlsx_path).exists():
                Path(xlsx_path).unlink()
            if temp_db_path.exists():
                temp_db_path.unlink()
        except Exception as e:
            print(f"Warning: Failed to delete temporary files: {e}")
        
        # Show summary
        summary = f"Report processed successfully!\n\n"
        summary += f"Jobs added: {len(jobs_added)}\n"
        summary += f"Jobs removed: {len(jobs_removed)}\n"
        summary += f"Jobs retained: {len(jobs_retained)}\n"
        summary += f"Total jobs: {len(merged_data)}"
        
        messagebox.showinfo("Success", summary)
        
    except Exception as e:
        messagebox.showerror("Processing Error", f"Failed to process report: {e}")
        def process_historic():
            """Process multiple historic E-reports to build log.json with historical data."""
            # Ask user to select directory containing historic reports
            folder_path = filedialog.askdirectory(
                title="Select folder containing historic E-reports (named DDMMYY)"
            )
            
            if not folder_path:
                return
            
            folder = Path(folder_path)
            
            # Find all .xls and .xlsx files in the folder
            report_files = list(folder.glob("*.xls")) + list(folder.glob("*.xlsx"))
            
            if not report_files:
                messagebox.showwarning("No Files", "No Excel files found in selected folder.")
                return
            
            # Parse filenames to get dates and sort chronologically
            dated_reports = []
            for file in report_files:
                filename = file.stem  # Get filename without extension
                
                # Try to parse as DDMMYY format
                try:
                    # Handle 6-digit format (DDMMYY)
                    if len(filename) == 6 and filename.isdigit():
                        day = int(filename[0:2])
                        month = int(filename[2:4])
                        year = int(filename[4:6]) + 2000  # Assume 2000s
                        
                        report_date = datetime(year, month, day).date()
                        dated_reports.append((report_date, file))
                except Exception:
                    # Skip files that don't match the naming convention
                    continue
            
            if not dated_reports:
                messagebox.showwarning("Invalid Files", "No files matching DDMMYY format found.")
                return
            
            # Sort by date
            dated_reports.sort(key=lambda x: x[0])
            
            # Confirm with user
            msg = f"Found {len(dated_reports)} historic reports.\n"
            msg += f"Date range: {dated_reports[0][0]} to {dated_reports[-1][0]}\n\n"
            msg += "This will process all reports and build historic log data.\n"
            msg += "Current database.json and log.json will be backed up.\n\n"
            msg += "Continue?"
            
            if not messagebox.askyesno("Confirm Processing", msg):
                return
            
            # Backup current files
            db_path = Path(__file__).parent / "database.json"
            log_path = Path(__file__).parent / "log.json"
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            try:
                if db_path.exists():
                    backup_db = db_path.with_name(f"database_backup_{timestamp}.json")
                    shutil.copy2(db_path, backup_db)
                
                if log_path.exists():
                    backup_log = log_path.with_name(f"log_backup_{timestamp}.json")
                    shutil.copy2(log_path, backup_log)
            except Exception as e:
                messagebox.showerror("Backup Error", f"Failed to backup files: {e}")
                return
            
            # Initialize empty database and log
            current_data = []
            log_data = []
            
            # Create temp directory
            temp_dir = Path(__file__).parent / "temp"
            temp_dir.mkdir(exist_ok=True)
            
            # Process each report in chronological order
            progress_window = tk.Toplevel(root)
            progress_window.title("Processing Historic Reports")
            progress_window.geometry("400x150")
            
            ttk.Label(progress_window, text="Processing reports...").pack(pady=10)
            progress_var = tk.StringVar(value="Starting...")
            progress_label = ttk.Label(progress_window, textvariable=progress_var)
            progress_label.pack(pady=10)
            
            progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate', maximum=len(dated_reports))
            progress_bar.pack(pady=10)
            
            root.update()
            
            for idx, (report_date, report_file) in enumerate(dated_reports):
                progress_var.set(f"Processing {report_date} ({idx + 1}/{len(dated_reports)})")
                progress_bar['value'] = idx + 1
                progress_window.update()
                
                try:
                    # Copy to temp location
                    temp_path = temp_dir / report_file.name
                    shutil.copy2(report_file, temp_path)
                    
                    # Convert if needed
                    if str(temp_path).endswith('.xls'):
                        xlsx_path = convert_report(temp_path)
                        if xlsx_path is None:
                            continue
                        temp_path = xlsx_path
                    
                    # Load and process workbook
                    wb = load_workbook(temp_path)
                    ws = wb.active
                    
                    # Delete column K and rename column L
                    ws.delete_cols(11)
                    ws.cell(row=1, column=11).value = 'Customer'
                    
                    # Get headers
                    headers = [cell.value for cell in ws[1]]
                    
                    # Process rows
                    new_data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] is None or len(str(row[0])) < 3:
                            continue
                        
                        row_data = {}
                        for header_idx, (header, value) in enumerate(zip(headers, row)):
                            if header_idx == 9 and isinstance(value, datetime):
                                row_data[header] = value.isoformat()
                            else:
                                row_data[header] = value
                        
                        row_data['Status'] = 'E'
                        row_data['Priority'] = False
                        row_data['Last Update'] = datetime.combine(report_date, datetime.min.time()).isoformat()
                        new_data.append(row_data)
                    
                    # Compare with previous data
                    old_job_nos = {entry.get('Job No') for entry in current_data if entry.get('Job No')}
                    new_job_nos = {entry.get('Job No') for entry in new_data if entry.get('Job No')}
                    
                    jobs_removed = old_job_nos - new_job_nos
                    jobs_added = new_job_nos - old_job_nos
                    jobs_retained = old_job_nos & new_job_nos
                    
                    # Calculate turnaround for removed jobs
                    removed_jobs_details = []
                    for job_no in jobs_removed:
                        for entry in current_data:
                            if entry.get('Job No') == job_no:
                                rec_date = entry.get('Rec_Date', '')
                                turnaround_days = None
                                
                                if rec_date:
                                    try:
                                        if isinstance(rec_date, str):
                                            rec_date_obj = datetime.fromisoformat(rec_date).date()
                                        else:
                                            rec_date_obj = rec_date.date() if hasattr(rec_date, 'date') else rec_date
                                        
                                        turnaround_days = (report_date - rec_date_obj).days
                                    except Exception:
                                        pass
                                
                                removed_jobs_details.append({
                                    'job_no': job_no,
                                    'customer': entry.get('Customer', ''),
                                    'turnaround_days': turnaround_days
                                })
                                break
                    
                    # Get added job details
                    added_jobs_details = []
                    for job_no in jobs_added:
                        for entry in new_data:
                            if entry.get('Job No') == job_no:
                                added_jobs_details.append({
                                    'job_no': job_no,
                                    'customer': entry.get('Customer', '')
                                })
                                break
                    
                    # Create log entry
                    log_entry = {
                        'timestamp': datetime.combine(report_date, datetime.min.time()).isoformat(),
                        'date': report_date.strftime('%Y-%m-%d'),
                        'report_type': 'E',
                        'total_jobs_before': len(current_data),
                        'total_jobs_after': len(new_data),
                        'jobs_added': len(jobs_added),
                        'jobs_removed': len(jobs_removed),
                        'jobs_retained': len(jobs_retained),
                        'removed_jobs': removed_jobs_details,
                        'added_jobs': added_jobs_details
                    }
                    
                    log_data.append(log_entry)
                    
                    # Update current data (preserve metadata for retained jobs)
                    merged_data = []
                    for new_entry in new_data:
                        job_no = new_entry.get('Job No')
                        
                        existing_entry = None
                        for old_entry in current_data:
                            if old_entry.get('Job No') == job_no:
                                existing_entry = old_entry
                                break
                        
                        if existing_entry:
                            new_entry['Priority'] = existing_entry.get('Priority', False)
                            new_entry['Assigned'] = existing_entry.get('Assigned', '')
                            if existing_entry.get('Status') == new_entry.get('Status'):
                                new_entry['Last Update'] = existing_entry.get('Last Update', new_entry['Last Update'])
                        
                        merged_data.append(new_entry)
                    
                    current_data = merged_data
                    
                    # Clean up temp file
                    if temp_path.exists():
                        temp_path.unlink()
                
                except Exception as e:
                    messagebox.showerror("Processing Error", f"Failed to process {report_file.name}: {e}")
                    continue
            
            # Save final log and database
            try:
                with open(log_path, 'w') as f:
                    json.dump(log_data, f, indent=4)
                
                with open(db_path, 'w') as f:
                    json.dump(current_data, f, indent=4)
                
                progress_window.destroy()
                
                # Calculate stats
                _calculate_stats()
                
                messagebox.showinfo("Success", f"Processed {len(dated_reports)} historic reports.\n\nLog and stats have been generated.")
            
            except Exception as e:
                progress_window.destroy()
                messagebox.showerror("Save Error", f"Failed to save results: {e}")

# Try to open maximized (cross-platform fallbacks)
try:
    root.state('zoomed')
except Exception:
    try:
        root.attributes('-zoomed', True)
    except Exception:
        pass

style = ttk.Style(root)
try:
    style.theme_use('default')
except Exception:
    try:
        style.theme_use(style.theme_names()[0])
    except Exception:
        pass

# Menu bar
menubar = tk.Menu(root)

# File menu
def _new_file():
    messagebox.showinfo("New", "IGNORE THIS.")

def _open_file():
    messagebox.showinfo("Load Legacy Priorities", "Load saved priorities file and update current joblist with this data.")

def _save_file():
    messagebox.showinfo("Save Priorities", "Save current priorities (not implemented).")

def _exit_app():
    if messagebox.askokcancel("Quit", "Do you want to quit Priorities?"):
        root.destroy()

file_menu = tk.Menu(menubar, tearoff=0)
file_menu.add_command(label="IGNORE", command=_new_file)
file_menu.add_command(label="Load Priorities", command=_open_file)
file_menu.add_command(label="Save Priorities", command=_save_file)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=_exit_app)
menubar.add_cascade(label="File", menu=file_menu)

# Edit menu
def _active_engineers():
    """Open popup to edit which engineers are active and save changes to settings.json."""
    # Load engineers from settings.json
    engineers = []
    current_settings = {}
    if SETTINGS.exists():
        try:
            with open(SETTINGS, 'r') as f:
                current_settings = json.load(f)
                engineers = current_settings.get('engineers', [])
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load settings: {e}")
            return
    
    if not engineers:
        messagebox.showwarning("No Engineers", "No engineers found in settings.json")
        return
    
    # Create popup
    popup = tk.Toplevel(root)
    popup.title("Active Engineers")
    popup.geometry("400x500")
    
    ttk.Label(popup, text="Toggle Active Engineers:", font=('TkDefaultFont', 12, 'bold')).pack(pady=10)
    ttk.Label(popup, text="Changes are saved immediately", font=('TkDefaultFont', 9, 'italic')).pack(pady=(0, 10))
    
    # Create scrollable frame
    canvas = tk.Canvas(popup)
    scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
    checkbox_frame = ttk.Frame(canvas)
    
    checkbox_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    
    canvas.create_window((0, 0), window=checkbox_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Dictionary to store checkbox variables
    checkboxes = {}
    
    def on_checkbox_change(eng_name):
        """Save settings immediately when a checkbox is toggled."""
        # Update engineer active status
        for eng in engineers:
            if eng['name'] == eng_name:
                eng['active'] = checkboxes[eng_name].get()
                break
        
        # Update settings with modified engineers list
        current_settings['engineers'] = engineers
        
        # Save to settings.json
        try:
            with open(SETTINGS, 'w') as f:
                json.dump(current_settings, f, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save settings: {e}")
    
    # Create checkbox for each engineer with immediate save on toggle
    for eng in engineers:
        var = tk.BooleanVar(value=eng.get('active', False))
        checkboxes[eng['name']] = var
        cb = ttk.Checkbutton(
            checkbox_frame, 
            text=eng['name'], 
            variable=var,
            command=lambda name=eng['name']: on_checkbox_change(name)
        )
        cb.pack(anchor="w", padx=20, pady=5)
    
    canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
    scrollbar.pack(side="right", fill="y", pady=10)
    
    # Close button at bottom
    button_frame = ttk.Frame(popup)
    button_frame.pack(fill="x", padx=10, pady=10)
    
    ttk.Button(button_frame, text="Close", command=popup.destroy).pack(side="left", padx=5)

edit_menu = tk.Menu(menubar, tearoff=0)
edit_menu.add_command(label="Active Engineers", command=_active_engineers)
menubar.add_cascade(label="Edit", menu=edit_menu)

# Utilities menu
def _open_settings():
    """Open settings dialog with tabs for Engineers, List Colours, and Statuses."""
    settings_window = tk.Toplevel(root)
    settings_window.title("Settings")
    settings_window.geometry("600x500")
    
    # Load current settings
    current_settings = {'engineers': [], 'list_colours': {}, 'statuses': [], 'late_threshold': 7}
    if SETTINGS.exists():
        try:
            with open(SETTINGS, 'r') as f:
                current_settings = json.load(f)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load settings: {e}")
    
    # Create notebook (tabs)
    notebook = ttk.Notebook(settings_window)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
    # --- Engineers Tab ---
    engineers_frame = ttk.Frame(notebook, padding=10)
    notebook.add(engineers_frame, text="Engineers")
    
    # List to store engineer data
    engineer_data = []
    
    def save_engineers():
        """Save only the engineers key to settings.json."""
        try:
            # Load current settings
            full_settings = {}
            if SETTINGS.exists():
                with open(SETTINGS, 'r') as f:
                    full_settings = json.load(f)
            
            # Update only engineers
            full_settings['engineers'] = engineer_data
            
            # Save back
            with open(SETTINGS, 'w') as f:
                json.dump(full_settings, f, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save engineers: {e}")
    
    def edit_engineer(eng_dict):
        """Open dialog to edit engineer details."""
        edit_window = tk.Toplevel(settings_window)
        edit_window.title(f"Edit Engineer: {eng_dict.get('name', 'New')}")
        edit_window.geometry("500x400")
        
        # Name
        ttk.Label(edit_window, text="Name:").grid(row=0, column=0, sticky="w", padx=10, pady=5)
        name_entry = ttk.Entry(edit_window, width=30)
        name_entry.grid(row=0, column=1, sticky="ew", padx=10, pady=5)
        name_entry.insert(0, eng_dict.get('name', ''))
        
        # Active status
        active_var = tk.BooleanVar(value=eng_dict.get('active', False))
        ttk.Checkbutton(edit_window, text="Active", variable=active_var).grid(row=1, column=0, columnspan=2, sticky="w", padx=10, pady=5)
        
        # Skills section
        ttk.Label(edit_window, text="Skills:").grid(row=2, column=0, sticky="nw", padx=10, pady=5)
        
        skills_frame = ttk.Frame(edit_window)
        skills_frame.grid(row=2, column=1, sticky="ew", padx=10, pady=5)
        
        skill_entries = []
        
        def add_skill(skill_text=""):
            skill_entry_frame = ttk.Frame(skills_frame)
            skill_entry_frame.pack(fill="x", pady=2)
            
            skill_entry = ttk.Entry(skill_entry_frame)
            skill_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
            skill_entry.insert(0, skill_text)
            
            def remove_skill():
                skill_entry_frame.destroy()
                skill_entries.remove((skill_entry_frame, skill_entry))
            
            remove_btn = ttk.Button(skill_entry_frame, text="Remove", command=remove_skill, width=8)
            remove_btn.pack(side="right")
            
            skill_entries.append((skill_entry_frame, skill_entry))
        
        # Add existing skills
        for skill in eng_dict.get('skills', []):
            add_skill(skill)
        
        ttk.Button(skills_frame, text="Add Skill", command=lambda: add_skill()).pack(pady=5)
        
        edit_window.grid_columnconfigure(1, weight=1)
        
        def save_engineer():
            name = name_entry.get().strip()
            if not name:
                messagebox.showwarning("Invalid", "Engineer name cannot be empty.")
                return
            
            skills = [entry.get().strip() for _, entry in skill_entries if entry.get().strip()]
            
            eng_dict['name'] = name
            eng_dict['active'] = active_var.get()
            eng_dict['skills'] = skills
            
            # Save immediately
            save_engineers()
            
            # Update display
            refresh_engineer_list()
            edit_window.destroy()
        
        button_frame = ttk.Frame(edit_window)
        button_frame.grid(row=3, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Save", command=save_engineer).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Cancel", command=edit_window.destroy).pack(side="left")
    
    def add_engineer():
        new_eng = {'name': 'New Engineer', 'active': False, 'skills': []}
        engineer_data.append(new_eng)
        save_engineers()
        refresh_engineer_list()
        edit_engineer(new_eng)
    
    def refresh_engineer_list():
        # Clear existing widgets
        for widget in engineers_list_frame.winfo_children():
            widget.destroy()
        
        # Rebuild list
        for eng in engineer_data:
            entry_frame = ttk.Frame(engineers_list_frame)
            entry_frame.pack(fill="x", pady=2)
            
            name_label = ttk.Label(entry_frame, text=eng.get('name', ''), width=30)
            name_label.pack(side="left", padx=(0, 5))
            
            ttk.Button(entry_frame, text="Edit", command=lambda e=eng: edit_engineer(e), width=8).pack(side="left", padx=2)
            
            def remove_this(e=eng):
                engineer_data.remove(e)
                save_engineers()
                refresh_engineer_list()
            
            ttk.Button(entry_frame, text="Remove", command=remove_this, width=8).pack(side="left")
    
    # Scrollable frame for engineers
    canvas = tk.Canvas(engineers_frame)
    scrollbar = ttk.Scrollbar(engineers_frame, orient="vertical", command=canvas.yview)
    engineers_list_frame = ttk.Frame(canvas)
    
    engineers_list_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=engineers_list_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Load existing engineers
    for eng in current_settings.get('engineers', []):
        engineer_data.append(eng.copy())
    
    refresh_engineer_list()
    
    # Add button for new engineers
    add_btn_frame = ttk.Frame(engineers_frame)
    add_btn_frame.pack(fill="x", pady=(10, 0))
    ttk.Button(add_btn_frame, text="Add Engineer", command=add_engineer).pack()
    
    # --- List Colours Tab ---
    colours_frame = ttk.Frame(notebook, padding=10)
    notebook.add(colours_frame, text="List Colours")
    
    colour_vars = {}
    
    def save_list_colours():
        """Save only the list_colours key to settings.json."""
        try:
            # Load current settings
            full_settings = {}
            if SETTINGS.exists():
                with open(SETTINGS, 'r') as f:
                    full_settings = json.load(f)
            
            # Update only list_colours
            full_settings['list_colours'] = colour_vars
            
            # Save back
            with open(SETTINGS, 'w') as f:
                json.dump(full_settings, f, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save list colours: {e}")
    
    def choose_colour(key, label_widget):
        settings_window.attributes('-topmost', False)
        color = colorchooser.askcolor(title=f"Choose colour for {key}", parent=settings_window)
        settings_window.attributes('-topmost', True)
        if color[1]:
            colour_vars[key] = color[1]
            label_widget.config(text=color[1], background=color[1])
            save_list_colours()
    
    for key, value in current_settings.get('list_colours', {}).items():
        row_frame = ttk.Frame(colours_frame)
        row_frame.pack(fill="x", pady=5)
        
        ttk.Label(row_frame, text=f"{key}:", width=20).pack(side="left")
        
        colour_vars[key] = value
        colour_label = tk.Label(row_frame, text=value, background=value, width=15, relief="solid", borderwidth=1)
        colour_label.pack(side="left", padx=5)
        
        ttk.Button(row_frame, text="Choose", command=lambda k=key, lbl=colour_label: choose_colour(k, lbl)).pack(side="left")
    
    # --- Statuses Tab ---
    statuses_frame = ttk.Frame(notebook, padding=10)
    notebook.add(statuses_frame, text="Statuses")
    
    status_list = []
    
    def save_statuses():
        """Save only the statuses key to settings.json."""
        try:
            # Load current settings
            full_settings = {}
            if SETTINGS.exists():
                with open(SETTINGS, 'r') as f:
                    full_settings = json.load(f)
            
            # Update only statuses
            full_settings['statuses'] = status_list
            
            # Save back
            with open(SETTINGS, 'w') as f:
                json.dump(full_settings, f, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save statuses: {e}")
    
    def refresh_status_list():
        for widget in status_listbox_frame.winfo_children():
            widget.destroy()
        
        for status in status_list:
            frame = ttk.Frame(status_listbox_frame)
            frame.pack(fill="x", pady=2)
            
            ttk.Label(frame, text=status, width=30).pack(side="left")
            
            def remove_this(s=status):
                status_list.remove(s)
                save_statuses()
                refresh_status_list()
            
            ttk.Button(frame, text="Remove", command=remove_this, width=8).pack(side="left")
    
    def add_status():
        new_status = status_entry.get().strip()
        if new_status and new_status not in status_list:
            status_list.append(new_status)
            status_entry.delete(0, tk.END)
            save_statuses()
            refresh_status_list()
    
    status_listbox_frame = ttk.Frame(statuses_frame)
    status_listbox_frame.pack(fill="both", expand=True, pady=(0, 10))
    
    # Load existing statuses
    for status in current_settings.get('statuses', []):
        status_list.append(status)
    
    refresh_status_list()
    
    add_frame = ttk.Frame(statuses_frame)
    add_frame.pack(fill="x")
    
    status_entry = ttk.Entry(add_frame)
    status_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
    
    ttk.Button(add_frame, text="Add Status", command=add_status).pack(side="left")
    
    # --- Late Threshold Tab ---
    threshold_frame = ttk.Frame(notebook, padding=10)
    notebook.add(threshold_frame, text="Late Threshold")
    
    ttk.Label(threshold_frame, text="Late threshold (days):").pack(pady=10)
    
    threshold_var = tk.IntVar(value=current_settings.get('late_threshold', 7))
    threshold_spinbox = ttk.Spinbox(threshold_frame, from_=1, to=365, textvariable=threshold_var, width=10)
    threshold_spinbox.pack()
    
    def save_threshold(*args):
        """Save only the late_threshold key to settings.json."""
        try:
            # Load current settings
            full_settings = {}
            if SETTINGS.exists():
                with open(SETTINGS, 'r') as f:
                    full_settings = json.load(f)
            
            # Update only late_threshold
            full_settings['late_threshold'] = threshold_var.get()
            
            # Save back
            with open(SETTINGS, 'w') as f:
                json.dump(full_settings, f, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save late threshold: {e}")
    
    # Save threshold when changed
    threshold_var.trace_add('write', save_threshold)
    
    # --- Close Button ---
    button_frame = ttk.Frame(settings_window)
    button_frame.pack(fill="x", padx=10, pady=(0, 10))
    
    ttk.Button(button_frame, text="Close", command=settings_window.destroy).pack(side="left", padx=5)

def import_reports():
    i = tk.Toplevel(root)
    i.title("Import Transmille Reports")
    i.geometry("600x400")
    
    # Map report types to status values
    report_types = {
        "E Reports:": "E",
        "F Reports:": "F",
        "L Reports:": "L",
        "N Reports:": "N",
        "A Reports:": "A"
    }
    
    selected_files = {}

    # Make columns behave nicely
    i.grid_columnconfigure(0, weight=1)
    i.grid_columnconfigure(1, weight=0)

    def _make_import_cmd(name, button, status):
        def _cmd():
            path = filedialog.askopenfilename(
                title=f"Import {name}", 
                filetypes=[("Excel File", "*.xls"), ("Excel File", "*.xlsx"), ("All files", "*.*")]
            )
            if path:
                selected_files[name] = path
                button.config(text="Processing...", state="disabled")
                i.update()
                
                # Copy file to temp folder before processing
                temp_dir = Path(__file__).parent / "temp"
                temp_dir.mkdir(exist_ok=True)
                temp_path = temp_dir / Path(path).name
                
                try:
                    shutil.copy2(path, temp_path)
                    process_report(temp_path, status)
                    i.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to import report: {e}")
                    button.config(text="Import", state="normal")
        return _cmd

    # ensure the Toplevel uses the same background as the app/theme
    bg = style.lookup('TLabel', 'background') or style.lookup('TFrame', 'background') or root.cget('background')
    if not bg:
        bg = root.cget('bg')
    i.configure(background=bg)

    for row, name in enumerate(report_types.keys()):
        lbl = ttk.Label(i, text=name)
        # force a background where ttk supports it (fallback safe)
        try:
            lbl.configure(background=bg)
        except Exception:
            pass
        lbl.grid(row=row, column=0, sticky="ew", padx=12, pady=8)

        btn = ttk.Button(i, text="Import", width=20)
        btn.config(command=_make_import_cmd(name, btn, report_types[name]))
        btn.grid(row=row, column=1, sticky="ew", padx=12, pady=8)

    # Close button at bottom
    ttk.Button(i, text="Close", command=i.destroy).grid(row=len(report_types), column=0, columnspan=2, pady=(16,8))

def _new_day():
    if messagebox.askyesno("New Day", "Start a new day? This will backup the current database and process a new E-report."):
        # Kill any widgets previously created by _show_list() if they exist
        for widget in root.pack_slaves():
            if isinstance(widget, ttk.Frame) and widget != toolbar:
                widget.destroy()
        
        d = tk.Toplevel(root)
        d.title("New Day Setup")
        d.geometry("400x500")
        
        # Load engineers from settings.json
        engineers = []
        current_settings = {}
        if SETTINGS.exists():
            try:
                with open(SETTINGS, 'r') as f:
                    current_settings = json.load(f)
                    engineers = current_settings.get('engineers', [])
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load settings: {e}")
                d.destroy()
                return
        
        # Dictionary to store checkbox variables
        checkboxes = {}
        
        # Frame for engineer checkboxes
        frame = ttk.Frame(d, padding=10)
        frame.pack(fill="both", expand=True)
        
        ttk.Label(frame, text="Select Active Engineers:").pack(anchor="w", pady=(0, 10))
        
        # Create checkbox for each engineer
        for eng in engineers:
            var = tk.BooleanVar(value=eng.get('active', False))
            checkboxes[eng['name']] = var
            ttk.Checkbutton(frame, text=eng['name'], variable=var).pack(anchor="w", pady=2)
        
        def _import_e_report():
            # Update engineer active status
            for eng in engineers:
                eng['active'] = checkboxes[eng['name']].get()
            
            # Update only the engineers in current_settings
            current_settings['engineers'] = engineers
            
            # Save updated settings
            try:
                with open(SETTINGS, 'w') as f:
                    json.dump(current_settings, f, indent=4)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save settings: {e}")
                return
            
            # Backup current database.json to history folder
            db_path = Path(__file__).parent / "database.json"
            history_dir = Path(__file__).parent / "history"
            history_dir.mkdir(exist_ok=True)
            
            if db_path.exists():
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_path = history_dir / f"database_{timestamp}.json"
                try:
                    shutil.copy2(db_path, backup_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to backup database: {e}")
                    return
            
            # Select E Report file
            path = filedialog.askopenfilename(
                title="Import E Report",
                filetypes=[("Excel File", "*.xls"), ("Excel File", "*.xlsx"), ("All files", "*.*")]
            )
            
            if path:
                # Copy to temp folder
                temp_dir = Path(__file__).parent / "temp"
                temp_dir.mkdir(exist_ok=True)
                dest = temp_dir / "e_report.xls"
                
                try:
                    shutil.copy2(path, dest)
                    process_report(dest, "E")
                    d.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to copy file: {e}")
        
        # Import button at bottom
        ttk.Button(d, text="Import E Report", command=_import_e_report).pack(pady=10)
        ttk.Button(d, text="Cancel", command=d.destroy).pack(pady=5)

def _print_report():
    """Open popup to select engineer and export their assigned jobs to Excel."""
    # Load active engineers from settings
    engineers = []
    if SETTINGS.exists():
        try:
            with open(SETTINGS, 'r') as f:
                settings = json.load(f)
                all_engineers = settings.get('engineers', [])
                engineers = [eng['name'] for eng in all_engineers if eng.get('active', False)]
        except Exception:
            pass
    
    if not engineers:
        messagebox.showwarning("No Engineers", "No active engineers found in settings.")
        return
    
    # Create popup
    popup = tk.Toplevel(root)
    popup.title("Export Report to Excel")
    popup.geometry("300x150")
    
    ttk.Label(popup, text="Select engineer:").pack(pady=10)
    
    engineer_var = tk.StringVar()
    engineer_dropdown = ttk.Combobox(popup, textvariable=engineer_var, values=engineers, state="readonly")
    engineer_dropdown.pack(pady=10)
    if engineers:
        engineer_dropdown.current(0)
    
    def generate_report():
        engineer_name = engineer_var.get()
        if not engineer_name:
            messagebox.showwarning("Invalid", "Please select an engineer.")
            return
        
        # Load database
        db_path = Path(__file__).parent / "database.json"
        if not db_path.exists():
            messagebox.showwarning("No Data", "database.json not found.")
            return
        
        try:
            with open(db_path, 'r') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load database: {e}")
            return
        
        # Filter for selected engineer
        assigned_jobs = [entry for entry in data if entry.get('Assigned') == engineer_name]
        
        if not assigned_jobs:
            messagebox.showinfo("No Jobs", f"No jobs assigned to {engineer_name}.")
            return
        
        # Sort by Rec_Date (oldest first)
        def get_rec_date(entry):
            rec_date = entry.get('Rec_Date', '')
            if rec_date:
                try:
                    if isinstance(rec_date, str):
                        return datetime.fromisoformat(rec_date)
                    return rec_date
                except Exception:
                    pass
            return datetime.min
        
        assigned_jobs.sort(key=get_rec_date)
        
        # Define columns for export
        columns = ['Job No', 'Serial_no', 'Manufacturer', 'Model_no', 'Description', 'Rec_Date', 'Customer']
        
        # Prepare data for DataFrame
        export_data = []
        for entry in assigned_jobs:
            row = {}
            for col in columns:
                value = entry.get(col, "")
                
                # Format Rec_Date
                if col == 'Rec_Date' and value:
                    try:
                        if isinstance(value, str):
                            dt = datetime.fromisoformat(value)
                        else:
                            dt = value
                        value = dt.strftime('%d/%m/%Y')
                    except Exception:
                        pass
                
                row[col] = value
            export_data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(export_data, columns=columns)
        
        # Ask user where to save
        popup.destroy()
        save_path = filedialog.asksaveasfilename(
            title=f"Save Report - {engineer_name}",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Report_{engineer_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if save_path:
            try:
                # Export to Excel
                df.to_excel(save_path, index=False, engine='openpyxl')
                messagebox.showinfo("Success", f"Report exported successfully to:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export report: {e}")
    
    ttk.Button(popup, text="Export", command=generate_report).pack(pady=5)
    ttk.Button(popup, text="Cancel", command=popup.destroy).pack(pady=5)

def _show_list():
    """Display database.json in a table format with selectable and colorable rows."""
    # Destroy existing widgets if they exist
    for widget in root.pack_slaves():
        if isinstance(widget, ttk.Frame) and widget != toolbar:
            widget.destroy()
    
    db_path = Path(__file__).parent / "database.json"
    
    if not db_path.exists():
        messagebox.showwarning("No Data", "database.json not found.")
        return
    
    try:
        with open(db_path, 'r') as f:
            data = json.load(f)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load database.json: {e}")
        return
    
    if not data:
        messagebox.showinfo("Empty", "No data in database.json")
        return
    
    # Load settings for colours and late_threshold
    list_colours = {'Priority': '#a85757', 'Late Dated': '#bd9582', 'Has Cert': '#57a857', 'Marked Complete': '#57a857'}
    late_threshold = 7  # Default 7 days
    if SETTINGS.exists():
        try:
            with open(SETTINGS, 'r') as f:
                settings = json.load(f)
                list_colours = settings.get('list_colours', list_colours)
                late_threshold = settings.get('late_threshold', late_threshold)
        except Exception:
            pass
    
    # Sort data: Priority items first, then by Rec_Date (oldest first)
    def sort_key(entry):
        is_priority = entry.get('Priority', False)
        rec_date = entry.get('Rec_Date', '')
        # Convert priority to sort first (False=1, True=0)
        priority_sort = 0 if is_priority else 1
        return (priority_sort, rec_date)
    
    data.sort(key=sort_key)
    
    # Define columns to display
    headers = ['Job No', 'Cert_No', 'Cust_Ref', 'Serial_no', 'Manufacturer', 
               'Model_no', 'Description', 'Rec_Date', 'Customer', 'Status', 'Assigned', 'Last Update']
    
    # Create search frame at top
    search_frame = ttk.Frame(root)
    search_frame.pack(side="top", fill="x", padx=10, pady=(10, 0))
    
    # Dictionary to hold search entry variables
    search_vars = {}
    
    # Get unique list of assigned engineers from data
    assigned_engineers = set()
    for entry in data:
        assigned = entry.get('Assigned', '')
        if assigned:
            assigned_engineers.add(assigned)
    assigned_engineers = sorted(list(assigned_engineers))
    
    for idx, col in enumerate(headers):
        search_frame.grid_columnconfigure(idx, weight=1)
        
        ttk.Label(search_frame, text=col, font=('TkDefaultFont', 8)).grid(row=0, column=idx, sticky="ew", padx=2)
        
        search_var = tk.StringVar()
        search_vars[col] = search_var
        
        # Use dropdown for Assigned column
        if col == 'Assigned':
            combo = ttk.Combobox(search_frame, textvariable=search_var, values=[''] + assigned_engineers, state="readonly")
            combo.grid(row=1, column=idx, sticky="ew", padx=2, pady=2)
            combo.bind('<<ComboboxSelected>>', lambda e: filter_data())
        else:
            entry = ttk.Entry(search_frame, textvariable=search_var)
            entry.grid(row=1, column=idx, sticky="ew", padx=2, pady=2)
            entry.bind('<KeyRelease>', lambda e: filter_data())
    
    # Create frame for table
    table_frame = ttk.Frame(root)
    table_frame.pack(side="top", fill="both", expand=True, padx=10, pady=10)
    
    # Create Treeview with scrollbars
    tree_scroll_y = ttk.Scrollbar(table_frame, orient="vertical")
    tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal")
    
    tree = ttk.Treeview(
        table_frame,
        columns=headers,
        show="headings",
        yscrollcommand=tree_scroll_y.set,
        xscrollcommand=tree_scroll_x.set,
        selectmode="extended"
    )
    
    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)
    
    # Configure tags for different colours
    tree.tag_configure('priority', background=list_colours.get('Priority', '#ff0000'))
    tree.tag_configure('late_dated', background=list_colours.get('Late Dated', '#ffcc00'))
    tree.tag_configure('has_cert', background=list_colours.get('Has Cert', '#00ff00'))
    tree.tag_configure('marked_complete', background=list_colours.get('Marked Complete', '#57a857'))
    
    # Configure column headings with initial width
    for col in headers:
        tree.heading(col, text=col)
        tree.column(col, width=100, anchor="w")
    
    # Get today's date for comparison
    today = datetime.now().date()
    
    # Variable to track current count
    count_var = tk.StringVar()
    
    def filter_data():
        """Filter and populate tree based on search criteria."""
        # Clear existing items
        for item in tree.get_children():
            tree.delete(item)
        
        # Get search filters
        filters = {col: search_vars[col].get().lower() for col in headers}
        
        # Filter and insert data
        visible_count = 0
        for entry in data:
            # Check if entry matches all filters
            match = True
            for col, search_text in filters.items():
                if search_text:
                    value = str(entry.get(col, "")).lower()
                    if search_text not in value:
                        match = False
                        break
            
            if not match:
                continue
            
            visible_count += 1
            values = []
            rec_date_obj = None
            
            for col in headers:
                value = entry.get(col, "")
                
                # Format Rec_Date as DD/MM/YYYY and store datetime object
                if col == 'Rec_Date' and value:
                    try:
                        if isinstance(value, str):
                            dt = datetime.fromisoformat(value)
                        else:
                            dt = value
                        rec_date_obj = dt.date()
                        value = dt.strftime('%d/%m/%Y')
                    except Exception:
                        pass
                
                # Format Last Update as DD/MM/YYYY HH:MM
                if col == 'Last Update' and value:
                    try:
                        if isinstance(value, str):
                            dt = datetime.fromisoformat(value)
                        else:
                            dt = value
                        value = dt.strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        pass
                
                values.append(value)
            
            # Determine tag based on priority rules
            tag = ()
            is_priority = entry.get('Priority', False)
            cert_no = entry.get('Cert_No', '')
            is_late = rec_date_obj and (today - rec_date_obj).days > late_threshold
            is_marked_complete = entry.get('Status', '') == 'N'
            
            if is_marked_complete:
                tag = ('marked_complete',)
            elif is_priority:
                tag = ('priority',)
            elif is_late:
                tag = ('late_dated',)
            elif cert_no:
                tag = ('has_cert',)
            
            tree.insert("", "end", values=values, tags=tag)
        
        # Update count display
        count_var.set(f"Current instruments: {visible_count}")
        
        # Auto-size columns to fit content
        for col in headers:
            max_width = len(col) * 10  # Header width
            for item in tree.get_children():
                item_text = str(tree.set(item, col))
                text_width = len(item_text) * 8
                max_width = max(max_width, text_width)
            tree.column(col, width=min(max_width + 20, 400))  # Cap at 400px
    
    # Initial population
    filter_data()
    
    # Pack scrollbars and treeview
    tree_scroll_y.pack(side="right", fill="y")
    tree_scroll_x.pack(side="bottom", fill="x")
    tree.pack(side="left", fill="both", expand=True)
    
    # Store reference for later color manipulation
    root._current_tree = tree
    
    # Options bar at bottom
    options_frame = ttk.Frame(root, height=60)
    options_frame.pack(side="bottom", fill="x", padx=10, pady=10)
    options_frame.pack_propagate(False)
    
    # Load icons for options buttons
    options_icon_dir = Path(__file__).parent / "icons"
    options_icons = {
        "Set Priority": "priority.png",
        "Change Status": "change_status.png",
        "Assign To...": "assign.png",
        "Unassign": "unassign.png",
        "Sub": "sub.png",
        "Refresh List": "refresh.png"
    }
    
    # Keep image references
    if not hasattr(root, "_options_img_refs"):
        root._options_img_refs = []
    
    def toggle_priority():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("No Selection", "Please select at least one item.")
            return
        
        for item in selected_items:
            job_no = tree.set(item, 'Job No')
            
            # Find and toggle in database
            for entry in data:
                if entry.get('Job No') == job_no:
                    entry['Priority'] = not entry.get('Priority', False)
                    entry['Last Update'] = datetime.now().isoformat()
                    break
        
        # Save updated database
        try:
            with open(db_path, 'w') as f:
                json.dump(data, f, indent=4)
            
            # Refresh the list
            _show_list()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update database: {e}")
    
    def change_status():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("No Selection", "Please select at least one item.")
            return
        
        # Load statuses from settings
        statuses = []
        if SETTINGS.exists():
            try:
                with open(SETTINGS, 'r') as f:
                    settings = json.load(f)
                    statuses = settings.get('statuses', [])
            except Exception:
                pass
        
        if not statuses:
            messagebox.showwarning("No Statuses", "No statuses found in settings.json")
            return
        
        # Create popup
        popup = tk.Toplevel(root)
        popup.title("Change Status")
        popup.geometry("300x150")
        
        ttk.Label(popup, text="Select new status:").pack(pady=10)
        
        status_var = tk.StringVar()
        status_dropdown = ttk.Combobox(popup, textvariable=status_var, values=statuses, state="readonly")
        status_dropdown.pack(pady=10)
        if statuses:
            status_dropdown.current(0)
        
        def apply_status():
            new_status = status_var.get()
            if not new_status:
                messagebox.showwarning("Invalid", "Please select a status.")
                return
            
            for item in selected_items:
                job_no = tree.set(item, 'Job No')
                
                # Update in database
                for entry in data:
                    if entry.get('Job No') == job_no:
                        entry['Status'] = new_status
                        entry['Last Update'] = datetime.now().isoformat()
                        break
            
            # Save updated database
            try:
                with open(db_path, 'w') as f:
                    json.dump(data, f, indent=4)
                
                popup.destroy()
                # Refresh the list
                _show_list()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update database: {e}")
        
        ttk.Button(popup, text="Apply", command=apply_status).pack(pady=5)
        ttk.Button(popup, text="Cancel", command=popup.destroy).pack(pady=5)
    
    def assign_to():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("No Selection", "Please select at least one item.")
            return
        
        # Load active engineers from settings
        engineers = []
        if SETTINGS.exists():
            try:
                with open(SETTINGS, 'r') as f:
                    settings = json.load(f)
                    all_engineers = settings.get('engineers', [])
                    engineers = [eng['name'] for eng in all_engineers if eng.get('active', False)]
            except Exception:
                pass
        
        if not engineers:
            messagebox.showwarning("No Engineers", "No active engineers found in settings.")
            return
        
        # Create popup
        popup = tk.Toplevel(root)
        popup.title("Assign To Engineer")
        popup.geometry("300x150")
        
        ttk.Label(popup, text="Select engineer:").pack(pady=10)
        
        engineer_var = tk.StringVar()
        engineer_dropdown = ttk.Combobox(popup, textvariable=engineer_var, values=engineers, state="readonly")
        engineer_dropdown.pack(pady=10)
        if engineers:
            engineer_dropdown.current(0)
        
        def apply_assignment():
            engineer_name = engineer_var.get()
            if not engineer_name:
                messagebox.showwarning("Invalid", "Please select an engineer.")
                return
            
            for item in selected_items:
                job_no = tree.set(item, 'Job No')
                
                # Update in database
                for entry in data:
                    if entry.get('Job No') == job_no:
                        entry['Assigned'] = engineer_name
                        entry['Last Update'] = datetime.now().isoformat()
                        break
            
            # Save updated database
            try:
                with open(db_path, 'w') as f:
                    json.dump(data, f, indent=4)
                
                popup.destroy()
                # Refresh the list
                _show_list()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update database: {e}")
        
        ttk.Button(popup, text="Assign", command=apply_assignment).pack(pady=5)
        ttk.Button(popup, text="Cancel", command=popup.destroy).pack(pady=5)
    
    def unassign():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("No Selection", "Please select at least one item.")
            return
        
        for item in selected_items:
            job_no = tree.set(item, 'Job No')
            
            # Remove Assigned field from database
            for entry in data:
                if entry.get('Job No') == job_no:
                    if 'Assigned' in entry:
                        del entry['Assigned']
                    entry['Last Update'] = datetime.now().isoformat()
                    break
        
        # Save updated database
        try:
            with open(db_path, 'w') as f:
                json.dump(data, f, indent=4)
            
            # Refresh the list
            _show_list()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update database: {e}")
    
    def assign_to_sub():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("No Selection", "Please select at least one item.")
            return
        
        for item in selected_items:
            job_no = tree.set(item, 'Job No')
            
            # Update in database
            for entry in data:
                if entry.get('Job No') == job_no:
                    entry['Assigned'] = 'Sub'
                    entry['Last Update'] = datetime.now().isoformat()
                    break
        
        # Save updated database
        try:
            with open(db_path, 'w') as f:
                json.dump(data, f, indent=4)
            
            # Refresh the list
            _show_list()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update database: {e}")
    
    def refresh_list():
        _show_list()
    
    # Map button labels to commands
    button_commands = {
        "Set Priority": toggle_priority,
        "Change Status": change_status,
        "Assign To...": assign_to,
        "Unassign": unassign,
        "Sub": assign_to_sub,
        "Refresh List": refresh_list
    }
    
    # Create buttons with icons
    for label, fname in options_icons.items():
        img = None
        fpath = options_icon_dir / fname
        try:
            if fpath.exists():
                img = tk.PhotoImage(file=str(fpath))
                root._options_img_refs.append(img)
        except Exception:
            img = None
        
        cmd = button_commands.get(label)
        
        if img:
            btn = ttk.Button(options_frame, text=label, image=img, compound="left", command=cmd)
        else:
            btn = ttk.Button(options_frame, text=label, command=cmd)
        
        btn.pack(side="left", padx=5)
    
    # Add count label on the right side
    count_label = ttk.Label(options_frame, textvariable=count_var, font=('TkDefaultFont', 10, 'bold'))
    count_label.pack(side="right", padx=10)

def _calculate_stats():
    """Calculate telemetrics and analytics from log.json and save to stats.json."""
    log_path = Path(__file__).parent / "log.json"
    stats_path = Path(__file__).parent / "stats.json"
    
    if not log_path.exists():
        messagebox.showwarning("No Data", "log.json not found.")
        return
    
    try:
        with open(log_path, 'r') as f:
            log_data = json.load(f)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load log.json: {e}")
        return
    
    if not log_data:
        messagebox.showinfo("Empty", "No data in log.json")
        return
    
    # Filter out weekend entries
    weekday_log_data = []
    for entry in log_data:
        entry_date = entry.get('date')
        if entry_date:
            try:
                date_obj = datetime.strptime(entry_date, '%Y-%m-%d')
                # Only include Monday-Friday (0-4)
                if date_obj.weekday() < 5:
                    weekday_log_data.append(entry)
            except Exception:
                # If date parsing fails, skip entry
                continue
    
    # Use filtered data for calculations
    log_data = weekday_log_data
    
    # Initialize statistics dictionary
    stats = {
        'generated_at': datetime.now().isoformat(),
        'total_updates': len(log_data),
        'date_range': {
            'first_update': None,
            'last_update': None,
            'weekdays_covered': 0
        },
        'overall_metrics': {
            'total_jobs_added': 0,
            'total_jobs_removed': 0,
            'total_jobs_retained': 0,
            'average_jobs_per_day': 0,
            'average_additions_per_day': 0,
            'average_removals_per_day': 0
        },
        'turnaround_analysis': {
            'completed_jobs': 0,
            'average_turnaround_days': 0,
            'median_turnaround_days': 0,
            'turnaround_distribution': {}
        },
        'by_report_type': {},
        'by_customer': {},
        'daily_summary': [],
        'trends': {
            'weekly_averages': [],
            'monthly_averages': []
        },
        'today': {
            'has_data': False,
            'date': None,
            'previous_date': None,
            'jobs_added': 0,
            'jobs_removed': 0,
            'net_change': 0,
            'jobs_before': 0,
            'jobs_after': 0,
            'turnaround_stats': {
                'completed_today': 0,
                'average_turnaround': 0
            }
        }
    }
    
    # Collect turnaround times and daily data
    all_turnaround_times = []
    customer_data = {}
    report_type_data = {}
    
    dates = []
    for entry in log_data:
        entry_date = entry.get('date')
        if entry_date:
            dates.append(entry_date)
        
        report_type = entry.get('report_type', 'Unknown')
        
        # Initialize report type stats
        if report_type not in report_type_data:
            report_type_data[report_type] = {
                'total_updates': 0,
                'total_added': 0,
                'total_removed': 0,
                'total_retained': 0
            }
        
        report_type_data[report_type]['total_updates'] += 1
        report_type_data[report_type]['total_added'] += entry.get('jobs_added', 0)
        report_type_data[report_type]['total_removed'] += entry.get('jobs_removed', 0)
        report_type_data[report_type]['total_retained'] += entry.get('jobs_retained', 0)
        
        # Overall metrics
        stats['overall_metrics']['total_jobs_added'] += entry.get('jobs_added', 0)
        stats['overall_metrics']['total_jobs_removed'] += entry.get('jobs_removed', 0)
        stats['overall_metrics']['total_jobs_retained'] += entry.get('jobs_retained', 0)
        
        # Process removed jobs for turnaround analysis
        for removed_job in entry.get('removed_jobs', []):
            turnaround_days = removed_job.get('turnaround_days')
            customer = removed_job.get('customer', 'Unknown')
            
            if turnaround_days is not None:
                all_turnaround_times.append(turnaround_days)
                
                # Customer statistics
                if customer not in customer_data:
                    customer_data[customer] = {
                        'jobs_completed': 0,
                        'total_turnaround_days': 0,
                        'turnaround_times': []
                    }
                
                customer_data[customer]['jobs_completed'] += 1
                customer_data[customer]['total_turnaround_days'] += turnaround_days
                customer_data[customer]['turnaround_times'].append(turnaround_days)
        
        # Daily summary
        stats['daily_summary'].append({
            'date': entry_date,
            'report_type': report_type,
            'jobs_before': entry.get('total_jobs_before', 0),
            'jobs_after': entry.get('total_jobs_after', 0),
            'jobs_added': entry.get('jobs_added', 0),
            'jobs_removed': entry.get('jobs_removed', 0),
            'net_change': entry.get('jobs_added', 0) - entry.get('jobs_removed', 0)
        })
    
    # Calculate "Today" stats (most recent update vs previous weekday)
    if len(log_data) >= 2:
        most_recent = log_data[-1]
        previous = log_data[-2]
        
        today_turnarounds = [job.get('turnaround_days') for job in most_recent.get('removed_jobs', []) if job.get('turnaround_days') is not None]
        
        stats['today'] = {
            'has_data': True,
            'date': most_recent.get('date'),
            'previous_date': previous.get('date'),
            'jobs_added': most_recent.get('jobs_added', 0),
            'jobs_removed': most_recent.get('jobs_removed', 0),
            'net_change': most_recent.get('jobs_added', 0) - most_recent.get('jobs_removed', 0),
            'jobs_before': most_recent.get('total_jobs_before', 0),
            'jobs_after': most_recent.get('total_jobs_after', 0),
            'turnaround_stats': {
                'completed_today': len(today_turnarounds),
                'average_turnaround': round(sum(today_turnarounds) / len(today_turnarounds), 2) if today_turnarounds else 0
            }
        }
    
    # Date range analysis (weekdays only)
    if dates:
        sorted_dates = sorted(dates)
        stats['date_range']['first_update'] = sorted_dates[0]
        stats['date_range']['last_update'] = sorted_dates[-1]
        
        try:
            first_date = datetime.strptime(sorted_dates[0], '%Y-%m-%d')
            last_date = datetime.strptime(sorted_dates[-1], '%Y-%m-%d')
            
            # Count weekdays between dates
            weekday_count = 0
            current = first_date
            while current <= last_date:
                if current.weekday() < 5:  # Monday=0, Friday=4
                    weekday_count += 1
                current += pd.Timedelta(days=1)
            
            stats['date_range']['weekdays_covered'] = weekday_count
            
            # Calculate averages based on weekdays only
            if weekday_count > 0:
                stats['overall_metrics']['average_jobs_per_day'] = round(
                    stats['overall_metrics']['total_jobs_retained'] / weekday_count, 2
                )
                stats['overall_metrics']['average_additions_per_day'] = round(
                    stats['overall_metrics']['total_jobs_added'] / weekday_count, 2
                )
                stats['overall_metrics']['average_removals_per_day'] = round(
                    stats['overall_metrics']['total_jobs_removed'] / weekday_count, 2
                )
        except Exception:
            pass
    
    # Turnaround analysis
    if all_turnaround_times:
        stats['turnaround_analysis']['completed_jobs'] = len(all_turnaround_times)
        stats['turnaround_analysis']['average_turnaround_days'] = round(
            sum(all_turnaround_times) / len(all_turnaround_times), 2
        )
        
        sorted_times = sorted(all_turnaround_times)
        mid = len(sorted_times) // 2
        if len(sorted_times) % 2 == 0:
            stats['turnaround_analysis']['median_turnaround_days'] = round(
                (sorted_times[mid - 1] + sorted_times[mid]) / 2, 2
            )
        else:
            stats['turnaround_analysis']['median_turnaround_days'] = sorted_times[mid]
        
        # Turnaround distribution (by week ranges)
        distribution = {
            '0-7 days': 0,
            '8-14 days': 0,
            '15-30 days': 0,
            '31-60 days': 0,
            '61-90 days': 0,
            '90+ days': 0
        }
        
        for days in all_turnaround_times:
            if days <= 7:
                distribution['0-7 days'] += 1
            elif days <= 14:
                distribution['8-14 days'] += 1
            elif days <= 30:
                distribution['15-30 days'] += 1
            elif days <= 60:
                distribution['31-60 days'] += 1
            elif days <= 90:
                distribution['61-90 days'] += 1
            else:
                distribution['90+ days'] += 1
        
        stats['turnaround_analysis']['turnaround_distribution'] = distribution
    
    # By report type
    stats['by_report_type'] = report_type_data
    
    # By customer
    for customer, data in customer_data.items():
        if data['turnaround_times']:
            avg_turnaround = round(data['total_turnaround_days'] / data['jobs_completed'], 2)
            sorted_times = sorted(data['turnaround_times'])
            mid = len(sorted_times) // 2
            
            if len(sorted_times) % 2 == 0:
                median = round((sorted_times[mid - 1] + sorted_times[mid]) / 2, 2)
            else:
                median = sorted_times[mid]
            
            stats['by_customer'][customer] = {
                'jobs_completed': data['jobs_completed'],
                'average_turnaround_days': avg_turnaround,
                'median_turnaround_days': median
            }
    
    # Weekly trends (weekdays only)
    weekly_data = {}
    for entry in stats['daily_summary']:
        try:
            entry_date = datetime.strptime(entry['date'], '%Y-%m-%d')
            week_key = entry_date.strftime('%Y-W%U')
            
            if week_key not in weekly_data:
                weekly_data[week_key] = {
                    'week': week_key,
                    'total_added': 0,
                    'total_removed': 0,
                    'weekdays': 0
                }
            
            weekly_data[week_key]['total_added'] += entry['jobs_added']
            weekly_data[week_key]['total_removed'] += entry['jobs_removed']
            weekly_data[week_key]['weekdays'] += 1
        except Exception:
            pass
    
    for week, data in weekly_data.items():
        stats['trends']['weekly_averages'].append({
            'week': week,
            'avg_added_per_day': round(data['total_added'] / data['weekdays'], 2) if data['weekdays'] > 0 else 0,
            'avg_removed_per_day': round(data['total_removed'] / data['weekdays'], 2) if data['weekdays'] > 0 else 0
        })
    
    # Monthly trends (weekdays only, daily figures)
    monthly_data = {}
    for entry in stats['daily_summary']:
        try:
            entry_date = datetime.strptime(entry['date'], '%Y-%m-%d')
            month_key = entry_date.strftime('%Y-%m')
            
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'month': month_key,
                    'daily_data': [],
                    'total_added': 0,
                    'total_removed': 0,
                    'weekdays': 0
                }
            
            monthly_data[month_key]['daily_data'].append({
                'date': entry['date'],
                'jobs_added': entry['jobs_added'],
                'jobs_removed': entry['jobs_removed'],
                'jobs_after': entry['jobs_after']
            })
            monthly_data[month_key]['total_added'] += entry['jobs_added']
            monthly_data[month_key]['total_removed'] += entry['jobs_removed']
            monthly_data[month_key]['weekdays'] += 1
        except Exception:
            pass
    
    for month, data in monthly_data.items():
        stats['trends']['monthly_averages'].append({
            'month': month,
            'daily_data': data['daily_data'],
            'avg_added_per_day': round(data['total_added'] / data['weekdays'], 2) if data['weekdays'] > 0 else 0,
            'avg_removed_per_day': round(data['total_removed'] / data['weekdays'], 2) if data['weekdays'] > 0 else 0,
            'total_added': data['total_added'],
            'total_removed': data['total_removed'],
            'weekdays': data['weekdays']
        })
    
    # Save statistics
    try:
        with open(stats_path, 'w') as f:
            json.dump(stats, f, indent=4)
        messagebox.showinfo("Success", f"Statistics calculated and saved to stats.json\n(Weekdays only: {stats['date_range']['weekdays_covered']} days)")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save stats.json: {e}")

def _show_status():
    """Display KPI analytics and statistics from stats.json in a graphical dashboard."""
    stats_path = Path(__file__).parent / "stats.json"
    
    # Always calculate latest statistics before showing
    _calculate_stats()
    
    # Check if calculation succeeded
    if not stats_path.exists():
        messagebox.showerror("Error", "Failed to generate statistics.")
        return
    
    # Load statistics
    try:
        with open(stats_path, 'r') as f:
            stats = json.load(f)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load stats.json: {e}")
        return
    
    # Create status window
    status_window = tk.Toplevel(root)
    status_window.title("Status Dashboard")
    status_window.geometry("1200x800")
    
    # Create notebook for different views
    notebook = ttk.Notebook(status_window)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
    # --- Today Tab ---
    today_frame = ttk.Frame(notebook, padding=10)
    notebook.add(today_frame, text="Today")
    
    today_data = stats.get('today', {})
    
    if today_data.get('has_data'):
        # Header
        ttk.Label(today_frame, text=f"Today's Update: {today_data.get('date', 'N/A')}", font=('TkDefaultFont', 14, 'bold')).pack(pady=(0, 5))
        ttk.Label(today_frame, text=f"Previous Update: {today_data.get('previous_date', 'N/A')}", font=('TkDefaultFont', 10)).pack(pady=(0, 10))
        
        # Key metrics
        metrics_frame = ttk.Frame(today_frame)
        metrics_frame.pack(fill="x", pady=10)
        
        today_metrics = [
            ("Jobs Before", today_data.get('jobs_before', 0)),
            ("Jobs After", today_data.get('jobs_after', 0)),
            ("Net Change", today_data.get('net_change', 0)),
            ("Jobs Added", today_data.get('jobs_added', 0)),
            ("Jobs Removed", today_data.get('jobs_removed', 0)),
            ("Completed Today", today_data.get('turnaround_stats', {}).get('completed_today', 0))
        ]
        
        for idx, (label, value) in enumerate(today_metrics):
            row = idx // 3
            col = idx % 3
            
            metric_frame = ttk.LabelFrame(metrics_frame, text=label, padding=10)
            metric_frame.grid(row=row, column=col, padx=10, pady=10, sticky="ew")
            
            value_label = ttk.Label(metric_frame, text=str(value), font=('TkDefaultFont', 18, 'bold'))
            value_label.pack()
        
        metrics_frame.grid_columnconfigure(0, weight=1)
        metrics_frame.grid_columnconfigure(1, weight=1)
        metrics_frame.grid_columnconfigure(2, weight=1)
        
        # Turnaround statistics
        ttk.Label(today_frame, text="Today's Turnaround Statistics", font=('TkDefaultFont', 12, 'bold')).pack(pady=(20, 10))
        
        ta_frame = ttk.Frame(today_frame)
        ta_frame.pack(fill="x", pady=10)
        
        ta_stats = today_data.get('turnaround_stats', {})
        
        metric_frame = ttk.LabelFrame(ta_frame, text="Average Turnaround", padding=10)
        metric_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        value_label = ttk.Label(metric_frame, text=f"{ta_stats.get('average_turnaround', 0)} days", font=('TkDefaultFont', 14, 'bold'))
        value_label.pack()
        
        ta_frame.grid_columnconfigure(0, weight=1)
    else:
        ttk.Label(today_frame, text="No data available for today", font=('TkDefaultFont', 14)).pack(pady=50)
    
    # --- Overview Tab ---
    overview_frame = ttk.Frame(notebook, padding=10)
    notebook.add(overview_frame, text="Overview")
    
    # Header
    ttk.Label(overview_frame, text="Overall Metrics (Weekdays Only)", font=('TkDefaultFont', 14, 'bold')).pack(pady=(0, 10))
    
    # Metrics grid
    metrics_frame = ttk.Frame(overview_frame)
    metrics_frame.pack(fill="x", pady=10)
    
    overall = stats.get('overall_metrics', {})
    date_range = stats.get('date_range', {})
    
    metrics = [
        ("Total Jobs Added", overall.get('total_jobs_added', 0)),
        ("Total Jobs Removed", overall.get('total_jobs_removed', 0)),
        ("Avg Jobs/Day", overall.get('average_jobs_per_day', 0)),
        ("Avg Additions/Day", overall.get('average_additions_per_day', 0)),
        ("Avg Removals/Day", overall.get('average_removals_per_day', 0)),
        ("Weekdays Covered", date_range.get('weekdays_covered', 0))
    ]
    
    for idx, (label, value) in enumerate(metrics):
        row = idx // 2
        col = idx % 2
        
        metric_frame = ttk.LabelFrame(metrics_frame, text=label, padding=10)
        metric_frame.grid(row=row, column=col, padx=10, pady=10, sticky="ew")
        
        value_label = ttk.Label(metric_frame, text=str(value), font=('TkDefaultFont', 18, 'bold'))
        value_label.pack()
    
    metrics_frame.grid_columnconfigure(0, weight=1)
    metrics_frame.grid_columnconfigure(1, weight=1)
    
    # Date range info
    date_info = f"Data Range: {date_range.get('first_update', 'N/A')} to {date_range.get('last_update', 'N/A')} (Weekdays only)"
    ttk.Label(overview_frame, text=date_info, font=('TkDefaultFont', 10)).pack(pady=10)
    
    # Daily trend graph
    ttk.Label(overview_frame, text="Daily Job Trends (Last 30 Weekdays)", font=('TkDefaultFont', 12, 'bold')).pack(pady=(20, 10))
    
    daily_summary = stats.get('daily_summary', [])[-30:]  # Last 30 entries
    if daily_summary:
        graph_canvas = tk.Canvas(overview_frame, width=900, height=250, bg='white', relief='solid', borderwidth=1)
        graph_canvas.pack(pady=10)
        
        # Extract data for graph
        dates = [entry.get('date', '') for entry in daily_summary]
        jobs_after = [entry.get('jobs_after', 0) for entry in daily_summary]
        jobs_added = [entry.get('jobs_added', 0) for entry in daily_summary]
        jobs_removed = [entry.get('jobs_removed', 0) for entry in daily_summary]
        
        if jobs_after:
            max_jobs = max(jobs_after)
            max_change = max(max(jobs_added) if jobs_added else 0, max(jobs_removed) if jobs_removed else 0)
            
            # Graph dimensions
            padding = 40
            width = 900 - 2 * padding
            height = 250 - 2 * padding
            
            # Draw axes
            graph_canvas.create_line(padding, padding + height, padding + width, padding + height, width=2)
            graph_canvas.create_line(padding, padding, padding, padding + height, width=2)
            
            # Plot total jobs line
            if len(jobs_after) > 1:
                for i in range(len(jobs_after) - 1):
                    x1 = padding + (i / (len(jobs_after) - 1)) * width
                    y1 = padding + height - (jobs_after[i] / max_jobs * height) if max_jobs > 0 else padding + height
                    x2 = padding + ((i + 1) / (len(jobs_after) - 1)) * width
                    y2 = padding + height - (jobs_after[i + 1] / max_jobs * height) if max_jobs > 0 else padding + height
                    graph_canvas.create_line(x1, y1, x2, y2, fill='blue', width=2)
            
            # Y-axis labels
            for i in range(6):
                y = padding + height - (i / 5) * height
                val = int((i / 5) * max_jobs)
                graph_canvas.create_text(padding - 10, y, text=str(val), anchor='e', font=('TkDefaultFont', 8))
            
            # Labels
            graph_canvas.create_text(padding + width / 2, padding + height + 25, text="Date (Weekdays)", font=('TkDefaultFont', 10, 'bold'))
            graph_canvas.create_text(15, padding + height / 2, text="Jobs", font=('TkDefaultFont', 10, 'bold'), angle=90)
    
    # --- Turnaround Tab ---
    turnaround_frame = ttk.Frame(notebook, padding=10)
    notebook.add(turnaround_frame, text="Turnaround Analysis")
    
    ttk.Label(turnaround_frame, text="Turnaround Statistics", font=('TkDefaultFont', 14, 'bold')).pack(pady=(0, 10))
    
    turnaround = stats.get('turnaround_analysis', {})
    
    ta_metrics_frame = ttk.Frame(turnaround_frame)
    ta_metrics_frame.pack(fill="x", pady=10)
    
    ta_metrics = [
        ("Completed Jobs", turnaround.get('completed_jobs', 0)),
        ("Average Days", turnaround.get('average_turnaround_days', 0)),
        ("Median Days", turnaround.get('median_turnaround_days', 0))
    ]
    
    for idx, (label, value) in enumerate(ta_metrics):
        metric_frame = ttk.LabelFrame(ta_metrics_frame, text=label, padding=10)
        metric_frame.grid(row=0, column=idx, padx=10, pady=10, sticky="ew")
        
        value_label = ttk.Label(metric_frame, text=str(value), font=('TkDefaultFont', 16, 'bold'))
        value_label.pack()
    
    ta_metrics_frame.grid_columnconfigure(0, weight=1)
    ta_metrics_frame.grid_columnconfigure(1, weight=1)
    ta_metrics_frame.grid_columnconfigure(2, weight=1)
    
    # Distribution bar chart
    ttk.Label(turnaround_frame, text="Turnaround Distribution", font=('TkDefaultFont', 12, 'bold')).pack(pady=(20, 10))
    
    distribution = turnaround.get('turnaround_distribution', {})
    if distribution:
        dist_canvas = tk.Canvas(turnaround_frame, width=900, height=300, bg='white', relief='solid', borderwidth=1)
        dist_canvas.pack(pady=10)
        
        ranges = list(distribution.keys())
        values = list(distribution.values())
        max_val = max(values) if values else 1
        
        padding = 50
        width = 900 - 2 * padding
        height = 300 - 2 * padding
        bar_width = width / len(ranges) * 0.8
        spacing = width / len(ranges)
        
        # Draw axes
        dist_canvas.create_line(padding, padding + height, padding + width, padding + height, width=2)
        dist_canvas.create_line(padding, padding, padding, padding + height, width=2)
        
        # Draw bars
        colors = ['#4CAF50', '#8BC34A', '#FFC107', '#FF9800', '#FF5722', '#F44336']
        for i, (range_label, count) in enumerate(zip(ranges, values)):
            x = padding + i * spacing + (spacing - bar_width) / 2
            bar_height = (count / max_val * height) if max_val > 0 else 0
            y = padding + height - bar_height
            
            color = colors[i % len(colors)]
            dist_canvas.create_rectangle(x, y, x + bar_width, padding + height, fill=color, outline='black')
            dist_canvas.create_text(x + bar_width / 2, y - 10, text=str(count), font=('TkDefaultFont', 10, 'bold'))
            dist_canvas.create_text(x + bar_width / 2, padding + height + 15, text=range_label, angle=45, anchor='ne', font=('TkDefaultFont', 8))
        
        # Y-axis labels
        for i in range(6):
            y = padding + height - (i / 5) * height
            val = int((i / 5) * max_val)
            dist_canvas.create_text(padding - 10, y, text=str(val), anchor='e', font=('TkDefaultFont', 8))
    
    # --- Customer Analysis Tab ---
    customer_frame = ttk.Frame(notebook, padding=10)
    notebook.add(customer_frame, text="By Customer")
    
    ttk.Label(customer_frame, text="Customer Statistics", font=('TkDefaultFont', 14, 'bold')).pack(pady=(0, 10))
    
    # Create scrollable frame
    canvas = tk.Canvas(customer_frame)
    scrollbar = ttk.Scrollbar(customer_frame, orient="vertical", command=canvas.yview)
    customer_list_frame = ttk.Frame(canvas)
    
    customer_list_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    
    canvas.create_window((0, 0), window=customer_list_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    by_customer = stats.get('by_customer', {})
    
    # Sort by jobs completed
    sorted_customers = sorted(by_customer.items(), key=lambda x: x[1].get('jobs_completed', 0), reverse=True)
    
    for customer, data in sorted_customers:
        cust_frame = ttk.LabelFrame(customer_list_frame, text=customer, padding=10)
        cust_frame.pack(fill="x", padx=10, pady=5)
        
        info = f"Jobs: {data.get('jobs_completed', 0)} | Avg: {data.get('average_turnaround_days', 0)} days | Median: {data.get('median_turnaround_days', 0)} days"
        ttk.Label(cust_frame, text=info).pack()
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # --- Trends Tab ---
    trends_frame = ttk.Frame(notebook, padding=10)
    notebook.add(trends_frame, text="Trends")
    
    ttk.Label(trends_frame, text="Monthly Trends (Daily Data - Weekdays Only)", font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 10))
    
    monthly_trends = stats.get('trends', {}).get('monthly_averages', [])
    
    # Dropdown to select month
    if monthly_trends:
        month_selection_frame = ttk.Frame(trends_frame)
        month_selection_frame.pack(pady=10)
        
        ttk.Label(month_selection_frame, text="Select Month:").pack(side="left", padx=5)
        
        month_options = [m.get('month', '') for m in monthly_trends]
        selected_month = tk.StringVar(value=month_options[-1] if month_options else "")
        
        month_dropdown = ttk.Combobox(month_selection_frame, textvariable=selected_month, values=month_options, state="readonly", width=15)
        month_dropdown.pack(side="left", padx=5)
        
        # Canvas for daily graph
        daily_canvas = tk.Canvas(trends_frame, width=900, height=300, bg='white', relief='solid', borderwidth=1)
        daily_canvas.pack(pady=10)
        
        def update_daily_graph(*args):
            month = selected_month.get()
            daily_canvas.delete("all")
            
            # Find selected month data
            month_data = None
            for m in monthly_trends:
                if m.get('month') == month:
                    month_data = m
                    break
            
            if not month_data or not month_data.get('daily_data'):
                return
            
            daily_data = month_data['daily_data']
            dates = [d.get('date', '') for d in daily_data]
            jobs_after = [d.get('jobs_after', 0) for d in daily_data]
            jobs_added = [d.get('jobs_added', 0) for d in daily_data]
            jobs_removed = [d.get('jobs_removed', 0) for d in daily_data]
            
            max_jobs = max(jobs_after) if jobs_after else 1
            
            padding = 50
            width = 900 - 2 * padding
            height = 300 - 2 * padding
            
            # Draw axes
            daily_canvas.create_line(padding, padding + height, padding + width, padding + height, width=2)
            daily_canvas.create_line(padding, padding, padding, padding + height, width=2)
            
            # Plot lines
            if len(dates) > 1:
                for i in range(len(dates) - 1):
                    x1 = padding + (i / (len(dates) - 1)) * width
                    x2 = padding + ((i + 1) / (len(dates) - 1)) * width
                    
                    # Total jobs line (blue)
                    y1 = padding + height - (jobs_after[i] / max_jobs * height) if max_jobs > 0 else padding + height
                    y2 = padding + height - (jobs_after[i + 1] / max_jobs * height) if max_jobs > 0 else padding + height
                    daily_canvas.create_line(x1, y1, x2, y2, fill='blue', width=2)
            
            # Y-axis labels
            for i in range(6):
                y = padding + height - (i / 5) * height
                val = int((i / 5) * max_jobs)
                daily_canvas.create_text(padding - 10, y, text=str(val), anchor='e', font=('TkDefaultFont', 8))
            
            # X-axis labels (show some dates)
            step = max(1, len(dates) // 10)
            for i in range(0, len(dates), step):
                x = padding + (i / (len(dates) - 1)) * width if len(dates) > 1 else padding
                date_str = dates[i][-5:]  # Show MM-DD
                daily_canvas.create_text(x, padding + height + 15, text=date_str, angle=45, anchor='ne', font=('TkDefaultFont', 7))
            
            # Title
            daily_canvas.create_text(padding + width / 2, 20, text=f"Daily Jobs for {month}", font=('TkDefaultFont', 11, 'bold'))
        
        month_dropdown.bind('<<ComboboxSelected>>', update_daily_graph)
        update_daily_graph()  # Initial display
    
    # Monthly summary details
    details_canvas = tk.Canvas(trends_frame)
    details_scrollbar = ttk.Scrollbar(trends_frame, orient="vertical", command=details_canvas.yview)
    trends_list_frame = ttk.Frame(details_canvas)
    
    trends_list_frame.bind("<Configure>", lambda e: details_canvas.configure(scrollregion=details_canvas.bbox("all")))
    
    details_canvas.create_window((0, 0), window=trends_list_frame, anchor="nw")
    details_canvas.configure(yscrollcommand=details_scrollbar.set)
    
    for month_data in monthly_trends:
        month_frame = ttk.LabelFrame(trends_list_frame, text=month_data.get('month', 'Unknown'), padding=10)
        month_frame.pack(fill="x", padx=10, pady=5)
        
        info = f"Weekdays: {month_data.get('weekdays', 0)} | Total Added: {month_data.get('total_added', 0)} | Total Removed: {month_data.get('total_removed', 0)}\n"
        info += f"Avg Added/Day: {month_data.get('avg_added_per_day', 0)} | Avg Removed/Day: {month_data.get('avg_removed_per_day', 0)}"
        ttk.Label(month_frame, text=info, justify="left").pack()
    
    details_canvas.pack(side="left", fill="both", expand=True)
    details_scrollbar.pack(side="right", fill="y")
    
    # Bottom buttons
    button_frame = ttk.Frame(status_window)
    button_frame.pack(fill="x", padx=10, pady=(0, 10))
    
    ttk.Button(button_frame, text="Recalculate Stats", command=lambda: [status_window.destroy(), _calculate_stats(), _show_status()]).pack(side="left", padx=5)
    ttk.Button(button_frame, text="Close", command=status_window.destroy).pack(side="left")


utilities_menu = tk.Menu(menubar, tearoff=0)
utilities_menu.add_command(label="New Day", command=_new_day)
utilities_menu.add_command(label="Import Reports", command=import_reports)
utilities_menu.add_command(label="Print", command=_print_report)
utilities_menu.add_separator()
utilities_menu.add_command(label="Show List", command=_show_list)
utilities_menu.add_command(label="Analytics", command=_show_status)
utilities_menu.add_separator()
utilities_menu.add_command(label="Settings", command=_open_settings)
menubar.add_cascade(label="Utilities", menu=utilities_menu)

# About menu
def _about():
    messagebox.showinfo("About", f"{SETTINGS_DATA.get('app_name', ' ')}\nVersion: {SETTINGS_DATA.get('version', ' ')}\n\nA Thru-Put handler for viewing/manipulating data from Transmille E-Reports.\n\n{SETTINGS_DATA.get('byline', '')}")

def _show_changelog():
    """Display changelog from changes.json in a popup with list and details view."""
    changes_path = Path(__file__).parent / "changes.json"
    
    if not changes_path.exists():
        messagebox.showwarning("No Changelog", "changes.json not found.")
        return
    
    try:
        with open(changes_path, 'r') as f:
            changelog_data = json.load(f)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load changes.json: {e}")
        return
    
    if not changelog_data:
        messagebox.showinfo("Empty", "No changelog entries found.")
        return
    
    # Create changelog window
    changelog_window = tk.Toplevel(root)
    changelog_window.title("Change Log")
    changelog_window.geometry("1000x600")
    
    # Main container with two panes
    main_frame = ttk.Frame(changelog_window)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Left pane - List of entries
    left_frame = ttk.Frame(main_frame)
    left_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
    
    ttk.Label(left_frame, text="Versions", font=('TkDefaultFont', 10, 'bold')).pack(pady=(0, 5))
    
    list_canvas = tk.Canvas(left_frame)
    list_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=list_canvas.yview)
    list_container = ttk.Frame(list_canvas)
    
    list_container.bind("<Configure>", lambda e: list_canvas.configure(scrollregion=list_canvas.bbox("all")))
    
    list_canvas.create_window((0, 0), window=list_container, anchor="nw", width=list_canvas.winfo_reqwidth())
    list_canvas.configure(yscrollcommand=list_scrollbar.set)
    
    # Update canvas window width when canvas is resized
    def on_canvas_configure(event):
        list_canvas.itemconfig(list_canvas.find_withtag("all")[0], width=event.width)
    
    list_canvas.bind('<Configure>', on_canvas_configure)
    
    list_canvas.pack(side="left", fill="both", expand=True)
    list_scrollbar.pack(side="right", fill="y")
    
    # Right pane - Details
    right_frame = ttk.Frame(main_frame)
    right_frame.pack(side="right", fill="both", expand=True)
    
    ttk.Label(right_frame, text="Details", font=('TkDefaultFont', 10, 'bold')).pack(pady=(0, 5))
    
    details_canvas = tk.Canvas(right_frame, relief='solid', borderwidth=1)
    details_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=details_canvas.yview)
    details_container = ttk.Frame(details_canvas)
    
    details_container.bind("<Configure>", lambda e: details_canvas.configure(scrollregion=details_canvas.bbox("all")))
    
    details_canvas_window = details_canvas.create_window((0, 0), window=details_container, anchor="nw")
    details_canvas.configure(yscrollcommand=details_scrollbar.set)
    
    # Update details canvas window width when canvas is resized
    def on_details_configure(event):
        details_canvas.itemconfig(details_canvas_window, width=event.width)
    
    details_canvas.bind('<Configure>', on_details_configure)
    
    # Enable mousewheel scrolling for details canvas
    def on_mousewheel(event):
        details_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    details_canvas.bind_all("<MouseWheel>", on_mousewheel)
    changelog_window.protocol("WM_DELETE_WINDOW", lambda: [details_canvas.unbind_all("<MouseWheel>"), changelog_window.destroy()])
    
    details_canvas.pack(side="left", fill="both", expand=True)
    details_scrollbar.pack(side="right", fill="y")
    
    # Variable to track selected entry
    selected_entry = tk.StringVar()
    
    def show_details(entry):
        """Display details of selected changelog entry."""
        # Clear existing details
        for widget in details_container.winfo_children():
            widget.destroy()
        
        # Display entry details
        ttk.Label(details_container, text=f"Version: {entry.get('Version Number', 'N/A')}", font=('TkDefaultFont', 12, 'bold')).pack(anchor="w", fill="x", padx=10, pady=(10, 5))
        ttk.Label(details_container, text=f"Date: {entry.get('Date', 'N/A')}", font=('TkDefaultFont', 10)).pack(anchor="w", fill="x", padx=10, pady=(0, 5))
        ttk.Label(details_container, text=f"Title: {entry.get('Title', 'N/A')}", font=('TkDefaultFont', 10, 'italic')).pack(anchor="w", fill="x", padx=10, pady=(0, 10))
        
        ttk.Separator(details_container, orient='horizontal').pack(fill='x', padx=10, pady=10)
        
        ttk.Label(details_container, text="Changes:", font=('TkDefaultFont', 10, 'bold')).pack(anchor="w", fill="x", padx=10, pady=(0, 5))
        
        changes_text = entry.get('Changes', '')
        if changes_text:
            # Split by newlines and display each line
            lines = changes_text.split('\n')
            for line in lines:
                line = line.strip()
                if line:
                    if line.startswith('-'):
                        # Bullet point
                        change_frame = ttk.Frame(details_container)
                        change_frame.pack(fill="x", padx=20, pady=2)
                        
                        ttk.Label(change_frame, text="•", font=('TkDefaultFont', 10)).pack(side="left", padx=(0, 5))
                        ttk.Label(change_frame, text=line[1:].strip(), font=('TkDefaultFont', 9), wraplength=400, justify="left").pack(side="left", fill="x", expand=True)
                    else:
                        # Regular text
                        ttk.Label(details_container, text=line, font=('TkDefaultFont', 9), wraplength=450, justify="left").pack(anchor="w", fill="x", padx=20, pady=2)
        else:
            ttk.Label(details_container, text="No changes listed.", font=('TkDefaultFont', 9, 'italic')).pack(anchor="w", fill="x", padx=20, pady=5)
    
    # Populate list with entries
    for entry in changelog_data:
        entry_frame = ttk.Frame(list_container, relief='solid', borderwidth=1, padding=5)
        entry_frame.pack(fill="x", padx=5, pady=2)
        
        version = entry.get('Version Number', 'N/A')
        date = entry.get('Date', 'N/A')
        title = entry.get('Title', 'N/A')
        
        ttk.Label(entry_frame, text=f"{version}", font=('TkDefaultFont', 9, 'bold')).pack(anchor="w", fill="x")
        ttk.Label(entry_frame, text=f"{date}", font=('TkDefaultFont', 8)).pack(anchor="w", fill="x")
        ttk.Label(entry_frame, text=f"{title}", font=('TkDefaultFont', 8, 'italic')).pack(anchor="w", fill="x")
        
        # Make entry clickable
        def make_click_handler(e=entry):
            return lambda event: show_details(e)
        
        for widget in [entry_frame] + list(entry_frame.winfo_children()):
            widget.bind("<Button-1>", make_click_handler(entry))
            widget.configure(cursor="hand2")
    
    # Show first entry by default
    if changelog_data:
        show_details(changelog_data[0])
    
    # Close button
    ttk.Button(changelog_window, text="Close", command=changelog_window.destroy).pack(pady=(0, 10))

def _show_help():
    """Display help tutorials from help.json in a popup with list and details view."""
    help_path = Path(__file__).parent / "help.json"
    
    if not help_path.exists():
        messagebox.showwarning("No Help", "help.json not found.")
        return
    
    try:
        with open(help_path, 'r') as f:
            help_data = json.load(f)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load help.json: {e}")
        return
    
    if not help_data:
        messagebox.showinfo("Empty", "No help tutorials found.")
        return
    
    # Create help window
    help_window = tk.Toplevel(root)
    help_window.title("Help & Tutorials")
    help_window.geometry("1200x600")
    
    # Main container with two panes
    main_frame = ttk.Frame(help_window)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Left pane - List of tutorials
    left_frame = ttk.Frame(main_frame)
    left_frame.pack(side="left", fill="both", expand=False, padx=(0, 5))
    
    ttk.Label(left_frame, text="Tutorials", font=('TkDefaultFont', 10, 'bold')).pack(pady=(0, 5))
    
    list_canvas = tk.Canvas(left_frame)
    list_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=list_canvas.yview)
    list_container = ttk.Frame(list_canvas)
    
    list_container.bind("<Configure>", lambda e: list_canvas.configure(scrollregion=list_canvas.bbox("all")))
    
    list_canvas.create_window((0, 0), window=list_container, anchor="nw", width=list_canvas.winfo_reqwidth())
    list_canvas.configure(yscrollcommand=list_scrollbar.set)
    
    # Update canvas window width when canvas is resized
    def on_canvas_configure(event):
        list_canvas.itemconfig(list_canvas.find_withtag("all")[0], width=event.width)
    
    list_canvas.bind('<Configure>', on_canvas_configure)
    
    list_canvas.pack(side="left", fill="both", expand=True)
    list_scrollbar.pack(side="right", fill="y")
    
    # Right pane - Details
    right_frame = ttk.Frame(main_frame)
    right_frame.pack(side="right", fill="both", expand=True)
    
    ttk.Label(right_frame, text="Tutorial Details", font=('TkDefaultFont', 10, 'bold')).pack(pady=(0, 5))
    
    details_canvas = tk.Canvas(right_frame, relief='solid', borderwidth=1)
    details_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=details_canvas.yview)
    details_container = ttk.Frame(details_canvas)
    
    details_container.bind("<Configure>", lambda e: details_canvas.configure(scrollregion=details_canvas.bbox("all")))
    
    details_canvas_window = details_canvas.create_window((0, 0), window=details_container, anchor="nw")
    details_canvas.configure(yscrollcommand=details_scrollbar.set)
    
    # Update details canvas window width when canvas is resized
    def on_details_configure(event):
        details_canvas.itemconfig(details_canvas_window, width=event.width)
    
    details_canvas.bind('<Configure>', on_details_configure)
    
    details_canvas.pack(side="left", fill="both", expand=True)
    details_scrollbar.pack(side="right", fill="y")
    
    def show_tutorial(tutorial):
        """Display details of selected tutorial."""
        # Clear existing details
        for widget in details_container.winfo_children():
            widget.destroy()
        
        # Display tutorial details
        ttk.Label(details_container, text=f"{tutorial.get('title', 'N/A')}", font=('TkDefaultFont', 12, 'bold')).pack(anchor="w", fill="x", padx=10, pady=(10, 5))
        ttk.Label(details_container, text=f"Date: {tutorial.get('date', 'N/A')}", font=('TkDefaultFont', 10)).pack(anchor="w", fill="x", padx=10, pady=(0, 10))
        
        ttk.Separator(details_container, orient='horizontal').pack(fill='x', padx=10, pady=10)
        
        ttk.Label(details_container, text="Content:", font=('TkDefaultFont', 10, 'bold')).pack(anchor="w", fill="x", padx=10, pady=(0, 5))
        
        content_text = tutorial.get('content', '')
        if content_text:
            # Split by newlines and display each line
            lines = content_text.split('\n')
            for line in lines:
                line_stripped = line.strip()
                if line_stripped:
                    # Check for indented lines (starting with tab or spaces)
                    indent_level = len(line) - len(line.lstrip())
                    
                    text_frame = ttk.Frame(details_container)
                    text_frame.pack(fill="x", padx=(20 + indent_level * 10), pady=2)
                    
                    ttk.Label(text_frame, text=line_stripped, font=('TkDefaultFont', 9), wraplength=400, justify="left").pack(anchor="w", fill="x")
        else:
            ttk.Label(details_container, text="No content available.", font=('TkDefaultFont', 9, 'italic')).pack(anchor="w", fill="x", padx=20, pady=5)
    
    # Populate list with tutorials
    for tutorial in help_data:
        entry_frame = ttk.Frame(list_container, relief='solid', borderwidth=1, padding=5)
        entry_frame.pack(fill="x", padx=5, pady=2)
        
        title = tutorial.get('title', 'N/A')
        date = tutorial.get('date', 'N/A')
        
        ttk.Label(entry_frame, text=f"{title}", font=('TkDefaultFont', 9, 'bold')).pack(anchor="w", fill="x")
        ttk.Label(entry_frame, text=f"{date}", font=('TkDefaultFont', 8, 'italic')).pack(anchor="w", fill="x")
        
        # Make entry clickable
        def make_click_handler(t=tutorial):
            return lambda event: show_tutorial(t)
        
        for widget in [entry_frame] + list(entry_frame.winfo_children()):
            widget.bind("<Button-1>", make_click_handler(tutorial))
            widget.configure(cursor="hand2")
    
    # Show first tutorial by default
    if help_data:
        show_tutorial(help_data[0])
    
    # Close button
    ttk.Button(help_window, text="Close", command=help_window.destroy).pack(pady=(0, 10))

about_menu = tk.Menu(menubar, tearoff=0)
about_menu.add_command(label="About", command=_about)
about_menu.add_command(label="Help", command=_show_help)
about_menu.add_command(label="Change Log", command=_show_changelog)
menubar.add_cascade(label="About", menu=about_menu)

root.config(menu=menubar)

# --- Toolbar with icon buttons ---


# toolbar frame
toolbar = ttk.Frame(root, padding=(6, 6, 6, 6))
toolbar.pack(side="top", fill="x")

# icon files expected in an "icons" directory next to this script
icon_dir = Path(__file__).parent / "icons"
icons = {
    "New Day": "new_day.png",
    "Import Report": "import.png",
    "Print": "print.png",
    "Show List": "list.png",
    "Analytics": "status.png",
}

# keep references so images are not garbage-collected
if not hasattr(root, "_img_refs"):
    root._img_refs = []

# mapping labels to commands
actions = {
    "New Day": _new_day,
    "Import Report": import_reports,
    "Print": _print_report,
    "Show List": _show_list,
    "Analytics": _show_status,
}

for label, fname in icons.items():
    img = None
    fpath = icon_dir / fname
    try:
        if fpath.exists():
            img = tk.PhotoImage(file=str(fpath))
            root._img_refs.append(img)
    except Exception:
        img = None

    cmd = actions.get(label, lambda: messagebox.showinfo(label, f"{label} (not implemented)."))

    if img:
        btn = ttk.Button(toolbar, text=label, image=img, compound="left", command=cmd)
    else:
        btn = ttk.Button(toolbar, text=label, command=cmd)

    btn.pack(side="left", padx=6, pady=2)

root.mainloop()